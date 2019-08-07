import cv2
import numpy as np
import os
import argparse
import datetime
import json
from openpyxl import Workbook,load_workbook

parser = argparse.ArgumentParser(description='Optik Okuyucu V.1.0')
parser.add_argument('--ogrencioptik', type=str, help='Ogrencilerin optiklerinin bulunduğu klasör...')
parser.add_argument('--cevapkagidi', type=str, help='Cevap Kağıtlarının olduğu klasör...')
parser.add_argument('--sonuckayit',type=str,help='Sonuçların txt olarak kaydedileceği yer...')
parser.add_argument('--sonucisim',type=str,help='Oluşturulacak excel dosyasının adı...')
parser.add_argument('--grup',type=int,help='Grup varmı yok mu ? ...')
parser.add_argument('--yanlisdogru',type=int,help='Kaç yanlış bir doğruyu götürsün? ...',default=100)
parser.add_argument('--puanlama',type=int,help='Sınav Kaç Üzerinden Hesaplansın?...',default=100)
args = parser.parse_args()

cevaplar=['A','B','C','D','E','-1','0']
harfler =['A', 'B', 'C', 'Ç', 'D', 'E', 'F', 'G', 'Ğ', 'H', 'I', 'İ', 'J', 'K', 'L', 'M', 'N', 'O', 'Ö', 'P', 'R', 'S', 'Ş', 'T', 'U', 'Ü', 'V', 'Y', 'Z', ' ']
cevap_kagidi = {}
cevap_kagidi_dict = {}
ogrenci_cevap = {}

class questions_number:
    soru_sayisi_static = 0

#ÖĞRENCİ KISMI
def alan_crop(image):
    ogrenci_cevap.clear()
    image = cv2.imread(image)
    image = cv2.resize(image,(600,800))
    isim_img = image[119:464,265:532]
    no_image = image[335:460, 71:192]
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (5, 5), 0)
    thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]
    if(args.grup == 1):
         grup = thresh[216:238, 130:253]
         ret = grup_belirle(grup)
    else:
        ret = 0
    ogrenci_cevap.setdefault('grup', ret)
    numara = ogrenci_no(no_image)
    isim = adi_soyadi(isim_img)
    artis = 80
    m = questions_number
    for i in range(0,m.soru_sayisi_static):
        cevap_1 = image[483:709,artis:artis+60]
        artis += 98
        ret=cevap_kontrol_krop(cevap_1,i)
    return ret,numara,isim

def ogrenci_no(image):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (7, 7), 0)
    thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]
    numara = ""
    artis = 1
    for i in range(0, 10):
        kolon = thresh[0:124,artis:artis+11]
        artis +=12
        sayi = sayi_cikar(kolon)
        numara += sayi
    return numara



def adi_soyadi(image):
    image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(image, (7,7), 0)
    thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]
    artis = 1
    isim = ""
    for i in range(0, 22):
        cevap = thresh[0:342,artis:artis+11]
        harf = harf_cikar(cevap)
        isim +=harf
        artis += 12
    return isim


def cevap_kontrol_krop(image,z):
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        blur = cv2.GaussianBlur(gray, (5, 5), 0)
        thresh = cv2.threshold(blur, 0, 255,cv2.THRESH_BINARY_INV| cv2.THRESH_OTSU)[1]
        artis = 0
        artis2 = 10
        if np.sum(thresh == 255) > 4000:
            for i in range(20):
                ogrenci_cevap.setdefault(z * 20 + i, cevaplar[6])
        else:
            for i in range(0,20):
                cevap = thresh[artis:artis2, 2:62]
                ret = secenekler(cevap)
                ogrenci_cevap.setdefault(z*20+i,ret)
                artis2 += 11
                artis += 11
                if i % 4 == 0:
                    artis += 1
                    artis2 += 1
        return ogrenci_cevap

#ÖĞRETMEN KISMI

def soru_sayisi(image):
    artis = 0
    artis2 = 10
    bos = 0
    for j in range(0, 20):
        cevap = image[artis:artis2, 2:62]
        ret = secenekler(cevap)
        if ret == '0':
            bos +=1
        artis2 += 11
        artis += 11
        if j % 4 == 0:
            artis += 1
            artis2 += 1
    if bos >= 10:
        return 0
    else:
        return 1

def cevap_kagidi_tanimla(img):
    cevap_kagidi_dict.clear()
    cevap_kagidi.clear()
    image = cv2.imread(img)
    image = cv2.resize(image, (600, 800))
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (5, 5), 0)
    thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]
    artis = 80
    if (args.grup == 1):
        grup = thresh[216:238,130:253]
        ret = grup_belirle(grup)
        f = open(str(ret) + '.txt', "w+")
    else:
        ret = 0
        f = open('cevaplar.txt','w+')
    f.close()
    for i in range(0, 5):
        cevap_1 = thresh[483:709, artis:artis + 60]
        sayi_ret = soru_sayisi(cevap_1)
        if sayi_ret == 0:
            m = questions_number
            m.soru_sayisi_static = i
            break
        artis += 98
        cevap_kagidi_siklar(cevap_1,i)
    if args.grup ==1:
        file1 = open(str(ret)+'.txt', 'a')
    else:
        file1 = open('cevaplar.txt','a')
    file1.write(json.dumps(cevap_kagidi_dict))
    file1.write('\n')
    file1.close()


def cevap_kagidi_siklar(image,i):
    artis = 0
    artis2 = 10
    for j in range(0, 20):
        cevap = image[artis:artis2, 2:62]
        ret = secenekler(cevap)
        if ret != str(-1):
            cevap_kagidi_dict.setdefault(i*20+j,ret)
        else:
            ret = hata_kontrol_cevap(cevap)
            key = str(i * 20 + j)
            cevap_kagidi_dict.setdefault(key, [])
            for x in ret:
                cevap_kagidi_dict[key].append(x)
        artis2 += 11
        artis += 11
        if j % 4 == 0:
            artis += 1
            artis2 += 1

def hata_kontrol_cevap(image):
    array = []
    artis = 0
    secenek = 0
    for j in range(0, 5):
        siklar = image[0:9, artis:artis + 11]
        artis += 12
        bps = np.sum(siklar == 255)
        if bps >= 40:
            secenek = j
            array.append(cevaplar[secenek])
    return array

#ORTAK KISIM


def sayi_cikar(img):
    artis = 0
    for i in range(0, 11):
        sayi = img[artis:artis + 11, 0:11]
        if i % 4 == 0:
            artis += 1
        artis += 11
        bps = np.sum(sayi == 255)
        if bps > 40 and i == 10:
            return '-'
        elif bps >40 and i<10:
            return str(i)

def secenekler(image):
    artis = 0
    array = []
    for j in range(0,5):
        siklar = image[0:9,artis:artis+12]
        artis +=12
        bps = np.sum(siklar == 255)
        if bps >= 30:
            array.append(cevaplar[j])
    if len(array) == 0:
        return cevaplar[6]
    elif len(array) == 1:
        return array[0]
    else:
        return cevaplar[5]

def harf_cikar(image):
    artis = 2
    for i in range(0,30):
        harf = image[artis:artis + 11,0:11]
        artis += 11
        if i%4 ==0:
            artis +=1
        bps = np.sum(harf == 255)
        if bps>50:
            return harfler[i]
        elif i==29 and bps<50:
            return harfler[-1]


def grup_belirle(image):
    if args.grup == 1:
        artis = 1
        secenek = -1
        for i in range(0, 4):
            grup = image[1:21,artis:artis+27]
            bps = np.sum(grup==255)
            artis += 25
            if bps>=60 and bps<=95:
                secenek = i
                break
        return cevaplar[secenek]
    else:
        return 0


for root, dirs, files in os.walk(args.cevapkagidi, topdown=False):
   for name in files:
       cevap_kagidi_tanimla(os.path.join(root,name))



def ogrenci_cevaplar(isim, numara, ogrenci_cevap, sira):
    if ogrenci_cevap['grup'] == 0:
        with open('cevaplar.txt') as answer_dict:
            cevap_kagidi = json.load(answer_dict)
    else:
        with open(str(ogrenci_cevap['grup'])+'.txt') as answer_dict:
            cevap_kagidi = json.load(answer_dict)
    ekleme = args.puanlama / len(cevap_kagidi)
    puan = 0
    bos = 0
    hatali = 0
    yanlis =0
    dogru = 0
    for i in range(len(cevap_kagidi)):
        cevap = list(cevap_kagidi[str(i)])
        if ogrenci_cevap[i] == '0':
           bos += 1
        else:
           if ogrenci_cevap[i] == '-1':
              hatali += 1
           else:
               if '0' in cevap:
                  continue
               elif ogrenci_cevap[i] in cevap:
                  puan += ekleme
                  dogru +=1
               else:
                   yanlis+=1
                   if yanlis % args.yanlisdogru == 0:
                       puan -=ekleme
    sheet1.cell(row=sira,column=1).value = numara
    sheet1.cell(row=sira,column=2).value = isim
    sheet1.cell(row=sira,column=3).value = puan
    sheet1.cell(row=sira,column= 4).value = bos
    sheet1.cell(row=sira,column= 5).value = hatali
    sheet1.cell(row=sira, column=6).value = yanlis
    sheet1.cell(row=sira, column=7).value = dogru

for root, dirs, files in os.walk(args.ogrencioptik, topdown=False):
    wb = Workbook()
    filepath = os.path.join(args.sonuckayit, args.sonucisim+'.xlsx')
    wb.save(filepath)
    wb = load_workbook(filepath)
    sheet1 = wb.active
    sheet1.title = str(datetime.date.today())
    sheet1.cell(row=1, column=1).value = 'NUMARA'
    sheet1.cell(row=1, column=2).value = 'İSİM'
    sheet1.cell(row=1, column=3).value = 'PUAN' + '( '+str(args.puanlama)+' uzerinden' + ')'
    sheet1.cell(row=1, column=4).value = 'Boş'
    sheet1.cell(row=1, column=5).value = 'Hatalı işaretleme'
    sheet1.cell(row=1, column=6).value = 'Yanlış Sayısı'
    sheet1.cell(row=1, column=7).value = 'Doğru Sayısı'
    sheet1.column_dimensions['A'].width = 19
    sheet1.column_dimensions['B'].width = 25
    sheet1.column_dimensions['C'].width = 30
    sheet1.column_dimensions['D'].width = 12
    sheet1.column_dimensions['E'].width = 20
    sheet1.column_dimensions['F'].width = 20
    sheet1.column_dimensions['G'].width = 18
    sira = 2
    for name in files:
        ogrenci_cevap.clear()
        answers,no,name=alan_crop(os.path.join(root, name))
        ogrenci_cevaplar(name,no,answers,sira)
        sira +=1
    wb.save(filepath)


